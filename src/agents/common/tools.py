import asyncio
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Annotated, Any

import pandas as pd
from langchain.tools import tool
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_core.tools import StructuredTool
from langchain_tavily import TavilySearch
from langgraph.types import interrupt
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from pydantic import BaseModel, Field, field_validator, model_validator

from src import config, graph_base, knowledge_base
from src.agents.common.models import load_chat_model
from src.utils import logger

search = TavilySearch(max_results=10)
search.metadata = {"name": "Tavily 网页搜索"}

EXPORT_BASE_DIR = Path(config.save_dir) / "exports"
EXPORT_BASE_DIR.mkdir(parents=True, exist_ok=True)

CYTHER_SYSTEM_PROMPT = (
"""您是一个基于Neo4j图数据库的Cypher查询生成专家。请根据给定的知识库结构，将自然语言问题转换为准确、可执行的Cypher查询语句。

    # 知识库定义
    ## 实体类型及关键属性
    以下是知识库中定义的实体类型及其关键属性（`field`属性可能包含多种标签，查询时请使用`label`属性）：
    - **组别**: 名称
    - **房**: 名称
    - **户**: 名称
    - **身份证号码**: 名称
    - **户籍地址**: 名称
    - **地址**: 名称
    - **联系方式**: 名称
    - **网格员**: 名称
    - **人口**: 包含`label`(标签，如“常住人口底册”、“重点优抚对象”等)和`名称`(姓名)在内的多个属性。
    - **产业**: 包含`label`(标签，如“小场所”、“种植户”等)和`名称`在内的多个属性。

    ## 实体关系
    实体间的关系定义如下，查询时请使用 `[]` 表示任意关系，并严格遵守方向：
    - `(人口)-[]->(身份证号码)`
    - `(人口)-[]->(联系方式)`
    - `(人口)-[]->(组别)`
    - `(人口)-[]->(房)`
    - `(人口)-[]->(户)`
    - `(人口)-[]->(户籍地址)`
    - `(人口)-[]->(网格员)`
    - `(产业)-[]->(组别)`
    - `(产业)-[]->(房)`
    - `(产业)-[]->(户)`
    - `(产业)-[]->(身份证号码)`
    - `(产业)-[]->(联系方式)`

    # 查询构建核心规则
    1.  **严格遵循知识库**：所有查询必须基于上述“实体类型”和“实体关系”构建。不得引入未定义的实体或关系。
    2.  **关系方向**：必须严格按照给定关系方向编写模式 `(实体1)-[]->(实体2)`。
    3.  **间接关系查询**：如果两个实体间没有直接关系，需通过中间实体串联。例如：`(a)-[]->(b)-[]->(c)`。
    4.  **实体匹配规则**：
        *   对`人口`或`产业`的`label`属性进行筛选时，必须使用 `CONTAINS` 匹配知识库`//`注释中提供的标签名称。
        *   对`名称`属性进行筛选时，使用 `CONTAINS` 或 `=`。
    5.  **数据类型与转换**：知识库中所有属性值为字符串。进行数学计算或比较时，请使用 `toInteger()`， `toFloat()` 等函数进行显式转换。
    6.  **输出控制**：
        *   当问题要求“列出”、“查询”、“有哪些”等详细信息时，直接 `RETURN` 实体本身（如 `RETURN p`）。
        *   当 `RETURN` 子句中显式列出的字段超过10个时，改为 `RETURN` 实体本身。
    7.  **分组列命名**：在进行分组统计（如分年龄段、分关系类型）时，用于分组的动态字段的别名必须命名为 `标题列`。
    8.  **语句规范性**：确保Cypher语法正确，变量引用一致，`AS`的别名中不得包含空格。
    9.  **查询范围**：仅生成数据查询和统计语句。不生成任何数据修改（如`CREATE`， `DELETE`）或模式操作语句。

    # 查询示例参考
    ### 基础统计
    1.  **统计各组别人口数量**：
        ```cypher
        MATCH (p:人口)-[]->(g:组别) RETURN g.名称 AS 组别, COUNT(p) AS 人口数量
        ```
    2.  **统计各组别的户数**：
        ```cypher
        MATCH (h:户)<-[]-(:人口)-[]->(g:组别) RETURN g.名称 AS 组别, COUNT(DISTINCT h) AS 户数量
        ```

    ### 条件筛选
    3.  **查询“重点优抚对象”**：
        ```cypher
        MATCH (p:人口)-[]->(g:组别) WHERE p.label CONTAINS ‘重点优抚对象’ RETURN g.名称 AS 组别, COUNT(p) AS 数量
        ```
    4.  **查询既是“退役军人”又是“志愿者”的人**：
        ```cypher
        MATCH (p:人口) WHERE p.label CONTAINS ‘退役军人’ AND p.label CONTAINS ‘社区志愿者’ RETURN p
        ```

    ### 数据质量与关联检查
    5.  **查找同一身份证对应不同姓名的人口**：
        ```cypher
        MATCH (id:身份证号码)<-[]-(p1:人口), (id)<-[]-(p2:人口) WHERE p1.名称 <> p2.名称 RETURN id.名称 AS 身份证号, COLLECT(DISTINCT p1.名称) AS 姓名列表
        ```
    6.  **验证身份证号码格式**：
        ```cypher
        MATCH (p:人口)-[]->(id:身份证号码) WHERE NOT id.名称 =~ ‘^[1-9]\d{5}(18|19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12][0-9]|3[01])\d{3}[0-9Xx]$’ RETURN p, id
        ```

    ### 复杂计算与分组
    7.  **按20岁间隔统计各年龄组人口**：
        ```cypher
        MATCH (p:人口)-[]->(g:组别) WITH g, p, toInteger(p.年龄) / 20 AS ageGroup RETURN g.名称 AS 组别, toString(ageGroup*20) + ‘-‘ + toString((ageGroup+1)*20-1) AS 标题列, COUNT(p) AS 人口数量 ORDER BY g.名称
        ```

    # 输出格式
    对于任何输入问题，**只输出最终生成的、符合上述所有规则的Cypher查询语句**，无需任何解释、推理过程或额外文本。

    **开始任务**
"""
)

_cypher_generator_model = None


def _sanitize_filename(name: str | None, fallback: str = "table") -> str:
    target = (name or fallback).strip()
    if not target:
        target = fallback
    return re.sub(r'[\\/:*?"<>|]+', "_", target)


def _build_export_path(file_name: str | None, suffix: str = "xlsx") -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized = _sanitize_filename(file_name, "table_export")
    export_dir = EXPORT_BASE_DIR / "excel"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir / f"{sanitized}_{timestamp}.{suffix}"


def _get_cypher_generator_model():
    global _cypher_generator_model
    if _cypher_generator_model is None:
        model_spec = getattr(config, "fast_model", None) or config.default_model
        _cypher_generator_model = load_chat_model(model_spec)
    return _cypher_generator_model


def _generate_cypher(query_text: str) -> str:
    model = _get_cypher_generator_model()
    messages = [
        SystemMessage(content=CYTHER_SYSTEM_PROMPT),
        HumanMessage(content=query_text),
    ]
    response = model.invoke(messages)
    cypher_query = getattr(response, "content", response)
    if isinstance(cypher_query, str):
        return _extract_cypher_from_output(cypher_query)
    if isinstance(cypher_query, list):
        combined = "".join([chunk.get("text", "") for chunk in cypher_query])
        return _extract_cypher_from_output(combined)
    return _extract_cypher_from_output(str(cypher_query))


def _extract_cypher_from_output(raw_text: str) -> str:
    """清洗模型返回内容，移除 Markdown 代码块，只保留 Cypher."""
    if not raw_text:
        raise ValueError("模型未返回任何内容，无法生成 Cypher")

    text = raw_text.strip()

    fence_pattern = re.compile(r"```(?:\w+)?\s*(.*?)```", re.IGNORECASE | re.DOTALL)
    fence_match = fence_pattern.search(text)
    if fence_match:
        text = fence_match.group(1).strip()

    if text.startswith("```") and text.endswith("```"):
        text = text[3:-3].strip()

    text = text.strip("`").strip()
    text = text.strip('"').strip("'").strip()

    if not text:
        raise ValueError(f"模型返回内容无法提取 Cypher：{raw_text!r}")
    return text


@tool(name_or_callable="计算器", description="可以对给定的2个数字选择进行 add, subtract, multiply, divide 运算")
def calculator(a: float, b: float, operation: str) -> float:
    try:
        if operation == "add":
            return a + b
        elif operation == "subtract":
            return a - b
        elif operation == "multiply":
            return a * b
        elif operation == "divide":
            if b == 0:
                raise ZeroDivisionError("除数不能为零")
            return a / b
        else:
            raise ValueError(f"不支持的运算类型: {operation}，仅支持 add, subtract, multiply, divide")
    except Exception as e:
        logger.error(f"Calculator error: {e}")
        raise


@tool(name_or_callable="人工审批工具(Debug)", description="请求人工审批工具，用于在执行重要操作前获得人类确认。")
def get_approved_user_goal(
    operation_description: str,
) -> dict:
    """
    请求人工审批，在执行重要操作前获得人类确认。

    Args:
        operation_description: 需要审批的操作描述，例如 "调用知识库工具"
    Returns:
        dict: 包含审批结果的字典，格式为 {"approved": bool, "message": str}
    """
    # 构建详细的中断信息
    interrupt_info = {
        "question": "是否批准以下操作？",
        "operation": operation_description,
    }

    # 触发人工审批
    is_approved = interrupt(interrupt_info)

    # 返回审批结果
    if is_approved:
        result = {
            "approved": True,
            "message": f"✅ 操作已批准：{operation_description}",
        }
        print(f"✅ 人工审批通过: {operation_description}")
    else:
        result = {
            "approved": False,
            "message": f"❌ 操作被拒绝：{operation_description}",
        }
        print(f"❌ 人工审批被拒绝: {operation_description}")

    return result


KG_QUERY_DESCRIPTION = """
使用这个工具可以查询知识图谱中包含的三元组信息。
关键词（query），使用可能帮助回答这个问题的关键词进行查询，不要直接使用用户的原始输入去查询。
"""


@tool(name_or_callable="查询知识图谱", description=KG_QUERY_DESCRIPTION)
def query_knowledge_graph(query: Annotated[str, "The keyword to query knowledge graph."]) -> Any:
    """使用这个工具可以查询知识图谱中包含的三元组信息。关键词（query），使用可能帮助回答这个问题的关键词进行查询，不要直接使用用户的原始输入去查询。"""
    try:
        logger.debug(f"Querying knowledge graph with: {query}")
        cypher_query = _generate_cypher(query)
        logger.debug(f"Generated Cypher: {cypher_query}")
        records = graph_base.run_cypher_query(cypher_query, return_format="records")
        logger.debug(f"Knowledge graph query returned {len(records) if isinstance(records, list) else 'N/A'} records")
        return {"records": records, "cypher": cypher_query}
    except Exception as e:
        logger.error(f"Knowledge graph query error: {e}, {traceback.format_exc()}")
        return f"知识图谱查询失败: {str(e)}"


def get_static_tools() -> list:
    """注册静态工具"""
    static_tools = [
        query_knowledge_graph,
        get_approved_user_goal,
        calculator,
        export_table_to_excel,
        auto_fill_excel_template,
    ]

    # 检查是否启用网页搜索
    if config.enable_web_search:
        static_tools.append(search)

    return static_tools


class KnowledgeRetrieverModel(BaseModel):
    query_text: str = Field(
        description=(
            "查询的关键词，查询的时候，应该尽量以可能帮助回答这个问题的关键词进行查询，不要直接使用用户的原始输入去查询。"
        )
    )
    operation: str = Field(
        default="search",
        description=(
            "操作类型：'search' 表示检索知识库内容，'get_mindmap' 表示获取知识库的思维导图结构。"
            "当用户询问知识库的整体结构、文件分类、知识架构时，使用 'get_mindmap'。"
            "当用户需要查询具体内容时，使用 'search'。"
        ),
    )


class TableExportModel(BaseModel):
    data: Annotated[str | list[dict[str, Any]] | list[list[Any]], Field(description="需要导出的表格数据，可以传入JSON字符串或直接传入列表")]  # type: ignore[type-arg]
    columns: list[str] | None = Field(
        default=None,
        description="当 data 为二维数组时指定列名；当 data 为对象列表且需要自定义列顺序时也可提供。",
    )
    file_name: str | None = Field(
        default=None,
        description="输出 Excel 文件名（不含扩展名），默认使用 table_export。",
    )
    sheet_name: str | None = Field(default="Sheet1", description="Excel 工作表名称，默认 Sheet1。")


class CellUpdate(BaseModel):
    cell: str = Field(..., description="需要写入的单元格（例如 B5）。")
    value: Any = Field(..., description="写入单元格的值。")

    @field_validator("cell")
    @classmethod
    def _validate_cell(cls, value: str) -> str:
        value = value.strip()
        if not re.fullmatch(r"[A-Za-z]+[1-9]\d*", value):
            raise ValueError("cell 必须是合法的单元格坐标，例如 A1、B12。")
        return value.upper()


class TableFillInstruction(BaseModel):
    start_cell: str = Field(..., description="数据块左上角单元格，如 B5。")
    data: list[list[Any]] = Field(..., description="需要写入的二维数组数据。")

    @field_validator("start_cell")
    @classmethod
    def _validate_start(cls, value: str) -> str:
        value = value.strip()
        if not re.fullmatch(r"[A-Za-z]+[1-9]\d*", value):
            raise ValueError("start_cell 必须是合法的单元格坐标。")
        return value.upper()

    @field_validator("data")
    @classmethod
    def _validate_data(cls, value: list[list[Any]]) -> list[list[Any]]:
        if not value:
            raise ValueError("data 不可为空。")
        return value


class ExcelAutoFillModel(BaseModel):
    template_path: str = Field(
        ...,
        description="Excel 模板文件路径，支持绝对路径或相对于项目根目录 / config.save_dir / uploads/chat_attachments目录的相对路径。",
    )
    output_file_name: str | None = Field(
        default=None,
        description="生成文件名称（可包含子目录），留空则自动生成，必须包含 .xlsx 扩展名或会自动补齐。",
    )
    sheet_name: str | None = Field(default=None, description="需要填充的工作表，未提供时默认使用第一个工作表。")
    cell_updates: list[CellUpdate] | None = Field(
        default=None,
        description="需要写入的零散单元格集合，例如 [{'cell': 'B2', 'value': '张三'}]。",
    )
    table_fill: TableFillInstruction | None = Field(
        default=None,
        description="如果需要按矩阵批量写入数据，可提供此字段来定义左上角及数据内容。",
    )
    header_row_data: list[dict[str, Any]] | None = Field(
        default=None,
        description="根据表头名称写入的行数据，每个元素为一行，如 [{'姓名': '张三', '本月收入': 1200}]。",
    )
    header_row_index: int | None = Field(
        default=1,
        description="表头所在的行号（默认第1行），用于基于列名定位列。需要为正整数。",
    )
    description: str | None = Field(
        default=None, description="可选的人类可读说明，便于记录在日志中，不影响程序执行。"
    )

    @field_validator("header_row_index")
    @classmethod
    def _validate_header_row_index(cls, value: int | None) -> int | None:
        if value is None:
            return value
        if value < 1:
            raise ValueError("header_row_index 必须为正整数。")
        return value

    @model_validator(mode="after")
    def check_payload(self):
        if not (self.cell_updates or self.table_fill or self.header_row_data):
            raise ValueError("必须至少提供 cell_updates、table_fill 或 header_row_data 其中之一。")
        return self


def _parse_table_data(data_input, columns: list[str] | None) -> pd.DataFrame:
    """将用户输入转换为 DataFrame."""
    if isinstance(data_input, str):
        try:
            payload = json.loads(data_input)
        except json.JSONDecodeError as exc:
            raise ValueError(f"无法解析的 JSON 数据: {exc}") from exc
    else:
        payload = data_input

    if isinstance(payload, list):
        if not payload:
            return pd.DataFrame(columns=columns or [])

        first_item = payload[0]
        if isinstance(first_item, dict):
            return pd.DataFrame(payload)

        if isinstance(first_item, (list, tuple)):
            if columns is None:
                raise ValueError("当 data 为二维数组时必须提供 columns。")
            return pd.DataFrame(payload, columns=columns)

    if isinstance(payload, dict):
        return pd.DataFrame([payload])

    raise ValueError("data 必须是对象列表、二维数组或可解析的 JSON 字符串。")


@tool(name_or_callable="表格导出为Excel", description="将结构化表格数据导出为本地 Excel 文件。", args_schema=TableExportModel)
def export_table_to_excel(data, columns: list[str] | None = None, file_name: str | None = None, sheet_name: str | None = "Sheet1") -> str:  # type: ignore[override]
    """
    将结构化表格数据导出为 Excel 文件。

    Args:
        data: JSON 字符串、对象列表或二维数组。
        columns: 当 data 为二维数组时所需的列名。
        file_name: 输出文件名（不含扩展名）。
        sheet_name: Excel 的工作表名称。
    """

    logger.debug(
        "Excel export requested",
        extra={
            "columns": columns,
            "file_name": file_name,
            "sheet_name": sheet_name,
            "data_type": type(data).__name__,
        },
    )

    df = _parse_table_data(data, columns)
    logger.debug(f"Parsed DataFrame shape: {df.shape}, columns: {list(df.columns)}")

    if df.empty:
        raise ValueError("表格数据为空，无法导出。")

    export_path = _build_export_path(file_name)
    os.makedirs(export_path.parent, exist_ok=True)
    logger.debug(f"Excel export path resolved: {export_path}")

    with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name or "Sheet1")

    logger.info(f"Excel 文件已导出: {export_path}")
    return str(export_path)


def _find_uploaded_attachment(file_name: str) -> Path | None:
    """
    在上传附件缓存目录中按原始文件名查找匹配文件。

    上传后的 Excel 会被保存为 `<file_id>_<original_name>.xlsx`，
    这里通过模糊匹配原始文件名来定位实际路径。
    """
    if not file_name:
        return None

    attachment_root = Path(config.save_dir) / "uploads" / "chat_attachments"
    if not attachment_root.exists():
        return None

    direct_path = (attachment_root / file_name).resolve()
    if direct_path.exists():
        return direct_path

    pattern = f"*_{file_name}"
    matches = sorted(
        attachment_root.glob(pattern),
        key=lambda p: p.stat().st_mtime if p.exists() else 0,
        reverse=True,
    )
    return matches[0].resolve() if matches else None


def _resolve_excel_path(path_str: str) -> Path:
    """解析 Excel 模板路径，支持多种相对目录以及上传附件。"""
    candidates: list[Path] = []
    raw_path = Path(path_str)
    if raw_path.is_absolute():
        candidates.append(raw_path)
    else:
        base_dirs = [
            Path.cwd(),
            Path(config.save_dir),
            Path(config.save_dir) / "uploads",
            Path.cwd() / "uploadFile",
        ]
        candidates.extend(base / raw_path for base in base_dirs)

    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved

    attachment_match = _find_uploaded_attachment(raw_path.name)
    if attachment_match and attachment_match.exists():
        logger.debug(f"Resolved Excel template via attachment cache: {attachment_match}")
        return attachment_match

    raise FileNotFoundError(f"未找到 Excel 模板文件: {path_str}")


def _build_autofill_export_path(source_path: Path, desired_name: str | None) -> Path:
    export_dir = EXPORT_BASE_DIR / "excel_autofill"
    export_dir.mkdir(parents=True, exist_ok=True)

    if desired_name:
        desired = Path(desired_name)
        if desired.suffix.lower() != ".xlsx":
            desired = desired.with_suffix(".xlsx")
        if desired.is_absolute():
            desired.parent.mkdir(parents=True, exist_ok=True)
            return desired
        return (export_dir / desired).resolve()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized = _sanitize_filename(source_path.stem, fallback="auto_fill")
    return (export_dir / f"{sanitized}_{timestamp}.xlsx").resolve()


def _set_cell_value(ws, row: int, column: int, value: Any) -> None:
    """
    安全地写入单元格：如果目标单元格属于合并区域，则自动拆分该区域后再写入，避免 MergedCell 只读错误。
    """
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        merged_range = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= column <= rng.max_col:
                merged_range = rng
                break
        if merged_range is not None:
            logger.warning(
                "Detected merged cell when writing value, unmerging to allow data fill",
                extra={"range": str(merged_range), "row": row, "column": column},
            )
            ws.unmerge_cells(str(merged_range))
            cell = ws.cell(row=row, column=column)
    cell.value = value


def _write_cell_updates(ws, updates: list[CellUpdate]) -> int:
    count = 0
    for update in updates:
        col_letters, row_str = coordinate_from_string(update.cell)
        row = int(row_str)
        column = column_index_from_string(col_letters)
        _set_cell_value(ws, row=row, column=column, value=update.value)
        count += 1
    return count


def _write_table_block(ws, table: TableFillInstruction) -> int:
    col_letters, row_str = coordinate_from_string(table.start_cell)
    start_row = int(row_str)
    start_col = column_index_from_string(col_letters)

    for row_offset, row_values in enumerate(table.data):
        for col_offset, value in enumerate(row_values):
            _set_cell_value(ws, row=start_row + row_offset, column=start_col + col_offset, value=value)
    return len(table.data)


def _normalize_header(value: Any) -> str:
    """标准化表头文本，移除括号内容和空白，便于模糊匹配。"""
    if value is None:
        return ""
    normalized = str(value).strip()
    if not normalized:
        return ""
    normalized = re.sub(r"[（(].*?[)）]", "", normalized)
    normalized = re.sub(r"[\s　:：_/\\-]+", "", normalized)
    return normalized


def _read_header_mapping(ws, header_row_index: int, *, strict: bool = True) -> tuple[dict[str, int], list[str]]:
    """读取表头行，返回列名到列号的映射及原始表头列表。"""
    headers: dict[str, int] = {}
    original_headers: list[str] = []
    header_row = ws[header_row_index]
    for cell in header_row:
        value = str(cell.value).strip() if cell.value is not None else ""
        if value:
            original_headers.append(value)
            headers[value] = cell.column
            normalized = _normalize_header(value)
            if normalized and normalized not in headers:
                headers[normalized] = cell.column
    if not headers and strict:
        raise ValueError(f"在第 {header_row_index} 行未检测到任何表头，请确认 header_row_index 是否正确。")
    return headers, original_headers


def _collect_desired_headers(rows: list[dict[str, Any]]) -> set[str]:
    desired: set[str] = set()
    for row in rows:
        for header in row.keys():
            normalized = _normalize_header(header)
            if normalized:
                desired.add(normalized)
    return desired


def _auto_detect_header_row(
    ws, desired_headers: set[str], max_scan_rows: int = 40
) -> tuple[int, dict[str, int], list[str]] | None:
    """根据期望字段自动扫描工作表的前几行，寻找最匹配的表头行。"""
    if not desired_headers:
        return None

    scan_limit = min(ws.max_row, max_scan_rows)
    best_row: int | None = None
    best_mapping: dict[str, int] = {}
    best_display: list[str] = []
    best_score = 0

    for row_idx in range(1, scan_limit + 1):
        mapping, display_headers = _read_header_mapping(ws, row_idx, strict=False)
        if not mapping:
            continue
        score = sum(1 for header in desired_headers if header in mapping)
        if score == 0:
            continue
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_mapping = mapping
            best_display = display_headers
            if score == len(desired_headers):
                break

    if best_row is None:
        return None
    return best_row, best_mapping, best_display


def _find_last_data_row(ws, start_row: int) -> int:
    """查找从 start_row 开始的最后一个包含数据的行，若没有则返回 start_row。"""
    last_row = start_row
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        if any(cell.value not in (None, "") for cell in row):
            last_row = row[0].row
    return last_row


def _append_rows_by_header(ws, headers: dict[str, int], rows: list[dict[str, Any]], start_row: int) -> int:
    """按照表头名称写入行数据，自动追加在最后一行之后。"""
    if not rows:
        return 0

    current_row = _find_last_data_row(ws, start_row) or start_row
    inserted = 0
    for row_data in rows:
        current_row += 1
        for header, value in row_data.items():
            header_key = _normalize_header(header)
            column_index = headers.get(header) or headers.get(header_key)
            if column_index is None:
                logger.warning(f"跳过未知表头 `{header}`，请确保字段名称与模板表头一致。")
                continue
            _set_cell_value(ws, row=current_row, column=column_index, value=value)
        inserted += 1
    return inserted


@tool(
    name_or_callable="Excel自动填表",
    description="根据结构化指令在 Excel 模板中写入数据，可支持指定单元格或批量区域写入，生成新的填报文件。",
    args_schema=ExcelAutoFillModel,
)
def auto_fill_excel_template(
    template_path: str,
    output_file_name: str | None = None,
    sheet_name: str | None = None,
    cell_updates: list[CellUpdate] | None = None,
    table_fill: TableFillInstruction | None = None,
    header_row_data: list[dict[str, Any]] | None = None,
    header_row_index: int | None = None,
    description: str | None = None,
) -> dict:
    """
    根据指令在 Excel 模板中填充值，并输出新的 Excel 文件路径。
    """

    logger.debug(
        "Auto fill Excel requested",
        extra={
            "template_path": template_path,
            "sheet_name": sheet_name,
            "has_cell_updates": bool(cell_updates),
            "has_table_fill": bool(table_fill),
        },
    )

    source_path = _resolve_excel_path(template_path)
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError("目前仅支持 .xlsx 文件。")

    workbook = load_workbook(source_path)

    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    cells_written = _write_cell_updates(worksheet, cell_updates or [])
    rows_written = _write_table_block(worksheet, table_fill) if table_fill else 0

    header_rows_written = 0
    template_headers_display: list[str] = []
    desired_header_names: list[str] = []
    if header_row_data:
        desired_headers = _collect_desired_headers(header_row_data)
        desired_header_names = sorted({_normalize_header(name) or str(name) for row in header_row_data for name in row.keys()})
        header_index = header_row_index or 1
        headers = None
        coverage = 0

        try:
            headers, template_headers_display = _read_header_mapping(worksheet, header_index)
            coverage = sum(1 for header in desired_headers if header in headers)
        except ValueError as exc:
            logger.warning(
                "Failed to read header row at index, attempting auto-detect",
                extra={"header_row_index": header_index, "error": str(exc)},
            )

        if (headers is None or coverage == 0) and desired_headers:
            detection = _auto_detect_header_row(worksheet, desired_headers)
            if detection:
                header_index, headers, template_headers_display = detection
                coverage = sum(1 for header in desired_headers if header in headers)
                logger.info(
                    "Auto-detected header row for Excel fill",
                    extra={"detected_row": header_index, "coverage": coverage, "desired_fields": list(desired_headers)},
                )

        if headers is None or coverage == 0:
            raise ValueError(
                "未能在模板中找到匹配的表头，请确认 header_row_index 是否正确，或检查模板列名是否与写入数据一致。"
            )

        header_rows_written = _append_rows_by_header(worksheet, headers, header_row_data, header_index)

    if cells_written == 0 and rows_written == 0 and header_rows_written == 0:
        raise ValueError("未提供有效的写入指令。")

    output_path = _build_autofill_export_path(source_path, output_file_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fallback_used = False
    try:
        workbook.save(output_path)
    except PermissionError as exc:  # noqa: PERF203
        logger.warning(
            "Failed to write Excel to desired path due to permission error, falling back to default location",
            extra={"output_path": str(output_path), "error": str(exc)},
        )
        output_path = _build_autofill_export_path(source_path, None)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        fallback_used = True

    logger.info(
        "Excel auto fill completed",
        extra={
            "source": str(source_path),
            "output": str(output_path),
            "sheet": worksheet.title,
            "cells_written": cells_written,
            "rows_written": rows_written,
            "header_rows_written": header_rows_written,
            "description": description,
            "fallback_used": fallback_used,
        },
    )

    return {
        "message": "Excel 自动填表完成",
        "output_path": str(output_path),
        "sheet": worksheet.title,
        "cells_written": cells_written,
        "rows_written": rows_written,
        "header_rows_written": header_rows_written,
        "description": description,
        "fallback_used": fallback_used,
        "template_headers": template_headers_display,
        "requested_headers": desired_header_names,
    }


def get_kb_based_tools() -> list:
    """获取所有知识库基于的工具"""
    # 获取所有知识库
    kb_tools = []
    retrievers = knowledge_base.get_retrievers()

    def _create_retriever_wrapper(db_id: str, retriever_info: dict[str, Any]):
        """创建检索器包装函数的工厂函数，避免闭包变量捕获问题"""

        async def async_retriever_wrapper(query_text: str, operation: str = "search") -> Any:
            """异步检索器包装函数，支持检索和获取思维导图"""

            # 获取思维导图
            if operation == "get_mindmap":
                try:
                    logger.debug(f"Getting mindmap for database {db_id}")

                    # 从知识库元数据中获取思维导图
                    if db_id not in knowledge_base.global_databases_meta:
                        return f"知识库 {retriever_info['name']} 不存在"

                    db_meta = knowledge_base.global_databases_meta[db_id]
                    mindmap_data = db_meta.get("mindmap")

                    if not mindmap_data:
                        return f"知识库 {retriever_info['name']} 还没有生成思维导图。"

                    # 将思维导图数据转换为文本格式，便于AI理解
                    def mindmap_to_text(node, level=0):
                        """递归将思维导图JSON转换为层级文本"""
                        indent = "  " * level
                        text = f"{indent}- {node.get('content', '')}\n"
                        for child in node.get("children", []):
                            text += mindmap_to_text(child, level + 1)
                        return text

                    mindmap_text = f"知识库 {retriever_info['name']} 的思维导图结构：\n\n"
                    mindmap_text += mindmap_to_text(mindmap_data)

                    logger.debug(f"Successfully retrieved mindmap for {db_id}")
                    return mindmap_text

                except Exception as e:
                    logger.error(f"Error getting mindmap for {db_id}: {e}")
                    return f"获取思维导图失败: {str(e)}"

            # 默认：检索知识库
            retriever = retriever_info["retriever"]
            try:
                logger.debug(f"Retrieving from database {db_id} with query: {query_text}")
                if asyncio.iscoroutinefunction(retriever):
                    result = await retriever(query_text)
                else:
                    result = retriever(query_text)
                logger.debug(f"Retrieved {len(result) if isinstance(result, list) else 'N/A'} results from {db_id}")
                return result
            except Exception as e:
                logger.error(f"Error in retriever {db_id}: {e}")
                return f"检索失败: {str(e)}"

        return async_retriever_wrapper

    for db_id, retrieve_info in retrievers.items():
        try:
            # 构建工具描述
            description = (
                f"使用 {retrieve_info['name']} 知识库的多功能工具。\n"
                f"知识库描述：{retrieve_info['description'] or '没有描述。'}\n\n"
                f"支持的操作：\n"
                f"1. 'search' - 检索知识库内容：根据关键词查询相关文档片段\n"
                f"2. 'get_mindmap' - 获取思维导图：查看知识库的整体结构和文件分类\n\n"
                f"使用建议：\n"
                f"- 需要查询具体内容时，使用 operation='search'\n"
                f"- 想了解知识库结构、文件分类时，使用 operation='get_mindmap'"
            )

            # 使用工厂函数创建检索器包装函数，避免闭包问题
            retriever_wrapper = _create_retriever_wrapper(db_id, retrieve_info)

            safename = retrieve_info["name"].replace(" ", "_")[:20]

            # 使用 StructuredTool.from_function 创建异步工具
            tool = StructuredTool.from_function(
                coroutine=retriever_wrapper,
                name=safename,
                description=description,
                args_schema=KnowledgeRetrieverModel,
                metadata=retrieve_info["metadata"] | {"tag": ["knowledgebase"]},
            )

            kb_tools.append(tool)
            # logger.debug(f"Successfully created tool {tool_id} for database {db_id}")

        except Exception as e:
            logger.error(f"Failed to create tool for database {db_id}: {e}, \n{traceback.format_exc()}")
            continue

    return kb_tools


def get_buildin_tools() -> list:
    """获取所有可运行的工具（给大模型使用）"""
    tools = []

    try:
        # 获取所有知识库基于的工具
        tools.extend(get_kb_based_tools())
        tools.extend(get_static_tools())

        from src.agents.common.toolkits.mysql.tools import get_mysql_tools

        tools.extend(get_mysql_tools())

    except Exception as e:
        logger.error(f"Failed to get knowledge base retrievers: {e}")

    return tools


def gen_tool_info(tools) -> list[dict[str, Any]]:
    """获取所有工具的信息（用于前端展示）"""
    tools_info = []

    try:
        # 获取注册的工具信息
        for tool_obj in tools:
            try:
                metadata = getattr(tool_obj, "metadata", {}) or {}
                info = {
                    "id": tool_obj.name,
                    "name": metadata.get("name", tool_obj.name),
                    "description": tool_obj.description,
                    "metadata": metadata,
                    "args": [],
                    # "is_async": is_async  # Include async information
                }

                if hasattr(tool_obj, "args_schema") and tool_obj.args_schema:
                    if isinstance(tool_obj.args_schema, dict):
                        schema = tool_obj.args_schema
                    else:
                        schema = tool_obj.args_schema.schema()

                    for arg_name, arg_info in schema.get("properties", {}).items():
                        info["args"].append(
                            {
                                "name": arg_name,
                                "type": arg_info.get("type", ""),
                                "description": arg_info.get("description", ""),
                            }
                        )

                tools_info.append(info)
                # logger.debug(f"Successfully processed tool info for {tool_obj.name}")

            except Exception as e:
                logger.error(
                    f"Failed to process tool {getattr(tool_obj, 'name', 'unknown')}: {e}\n{traceback.format_exc()}. "
                    f"Details: {dict(tool_obj.__dict__)}"
                )
                continue

    except Exception as e:
        logger.error(f"Failed to get tools info: {e}\n{traceback.format_exc()}")
        return []

    logger.info(f"Successfully extracted info for {len(tools_info)} tools")
    return tools_info
